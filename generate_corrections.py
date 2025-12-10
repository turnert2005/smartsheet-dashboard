"""
Generate Row-by-Row Corrections for Jan 13, 2026 Target
Creates an Excel file with specific corrections for each task
"""

import json
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def load_data():
    """Load current schedule"""
    with open('sheet_data_audit.json', 'r') as f:
        return json.load(f)


def parse_date(date_str):
    if not date_str:
        return None
    try:
        if 'T' in str(date_str):
            return datetime.strptime(str(date_str).split('T')[0], '%Y-%m-%d')
        return datetime.strptime(str(date_str), '%Y-%m-%d')
    except:
        return None


def calculate_corrections(tasks):
    """Calculate corrections needed for each task"""
    target = datetime(2026, 1, 13)
    original_baseline = datetime(2026, 1, 7)

    # Build task lookup
    task_by_row = {t['row_number']: t for t in tasks}

    corrections = []

    for task in tasks:
        row = task['row_number']
        end_date = parse_date(task.get('End Date'))
        baseline_finish = parse_date(task.get('Baseline Finish'))
        start_date = parse_date(task.get('Start Date'))
        baseline_start = parse_date(task.get('Baseline Start'))

        correction = {
            'Row': row,
            'Task': task['Tasks'][:50],
            'Assigned To': task.get('Assigned To', ''),
            'Status': task.get('Status', ''),
            'Health': task.get('Health', ''),

            # Current dates
            'Current Start': start_date.strftime('%Y-%m-%d') if start_date else '',
            'Current End': end_date.strftime('%Y-%m-%d') if end_date else '',
            'Current Baseline Start': baseline_start.strftime('%Y-%m-%d') if baseline_start else '',
            'Current Baseline Finish': baseline_finish.strftime('%Y-%m-%d') if baseline_finish else '',
            'Current Variance': task.get('Variance', ''),

            # Predecessor info
            'Current Predecessor': task.get('Predecessors', ''),

            # Corrections needed
            'Baseline Action': '',
            'New Baseline Start': '',
            'New Baseline Finish': '',
            'Predecessor Action': '',
            'New Predecessor': '',
            'Schedule Action': '',
            'Notes': ''
        }

        # Determine baseline correction
        if baseline_finish:
            if baseline_finish == original_baseline:
                # Tasks at Jan 7 baseline need update to Jan 13
                correction['Baseline Action'] = 'UPDATE TO JAN 13'
                correction['New Baseline Finish'] = '2026-01-13'
                if baseline_start:
                    # Shift start by same 6 days
                    new_start = baseline_start + timedelta(days=6)
                    correction['New Baseline Start'] = new_start.strftime('%Y-%m-%d')
            elif baseline_finish < original_baseline:
                # Earlier baselines - proportionally adjust
                days_before = (original_baseline - baseline_finish).days
                new_baseline = target - timedelta(days=days_before)
                correction['Baseline Action'] = 'PROPORTIONAL SHIFT'
                correction['New Baseline Finish'] = new_baseline.strftime('%Y-%m-%d')
                if baseline_start:
                    new_start = baseline_start + timedelta(days=6)
                    correction['New Baseline Start'] = new_start.strftime('%Y-%m-%d')

        # Determine if schedule compression needed
        if end_date and end_date > target:
            days_over = (end_date - target).days
            correction['Schedule Action'] = f'COMPRESS {days_over}d'
            correction['Notes'] = f'Currently {days_over} days past target'

        # Check predecessor issues
        pred = task.get('Predecessors')
        if pred:
            pred_row = int(pred.split('FS')[0].split('SS')[0].strip()) if 'FS' in pred or 'SS' in pred else None
            if pred_row and pred_row in task_by_row:
                pred_task = task_by_row[pred_row]
                pred_status = pred_task.get('Status', '')
                if pred_status == 'Complete' and task.get('Status') == 'Not Started':
                    correction['Predecessor Action'] = 'REVIEW - pred complete'
                    correction['Notes'] += ' | Predecessor done but task not started'

        # Check for Row 24 dependency (the bottleneck)
        if pred and '24' in pred:
            correction['Predecessor Action'] = 'CRITICAL - depends on IGT'
            correction['Notes'] += ' | Blocked by IGT Row 24'

        corrections.append(correction)

    return corrections


def create_excel_report(corrections):
    """Create Excel file with corrections"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Corrections Plan"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    # Headers
    headers = [
        'Row', 'Task', 'Assigned To', 'Status', 'Health',
        'Current End', 'Current Baseline Finish', 'Current Variance',
        'Baseline Action', 'New Baseline Finish',
        'Current Predecessor', 'Predecessor Action',
        'Schedule Action', 'Notes'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for row_idx, corr in enumerate(corrections, 2):
        ws.cell(row=row_idx, column=1, value=corr['Row'])
        ws.cell(row=row_idx, column=2, value=corr['Task'])
        ws.cell(row=row_idx, column=3, value=corr['Assigned To'])
        ws.cell(row=row_idx, column=4, value=corr['Status'])
        ws.cell(row=row_idx, column=5, value=corr['Health'])
        ws.cell(row=row_idx, column=6, value=corr['Current End'])
        ws.cell(row=row_idx, column=7, value=corr['Current Baseline Finish'])
        ws.cell(row=row_idx, column=8, value=corr['Current Variance'])
        ws.cell(row=row_idx, column=9, value=corr['Baseline Action'])
        ws.cell(row=row_idx, column=10, value=corr['New Baseline Finish'])
        ws.cell(row=row_idx, column=11, value=corr['Current Predecessor'])
        ws.cell(row=row_idx, column=12, value=corr['Predecessor Action'])
        ws.cell(row=row_idx, column=13, value=corr['Schedule Action'])
        ws.cell(row=row_idx, column=14, value=corr['Notes'])

        # Color coding
        health = corr['Health']
        if health == 'Red':
            ws.cell(row=row_idx, column=5).fill = red_fill
        elif health == 'Yellow':
            ws.cell(row=row_idx, column=5).fill = yellow_fill
        elif health == 'Green':
            ws.cell(row=row_idx, column=5).fill = green_fill

        # Highlight actions needed
        if corr['Baseline Action']:
            ws.cell(row=row_idx, column=9).fill = yellow_fill
        if corr['Predecessor Action']:
            ws.cell(row=row_idx, column=12).fill = yellow_fill
        if corr['Schedule Action']:
            ws.cell(row=row_idx, column=13).fill = red_fill

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 40

    # Freeze panes
    ws.freeze_panes = 'C2'

    # Add summary sheet
    ws2 = wb.create_sheet("Summary")

    summaries = [
        ["PHASE 2 AGENTIC VOICE - SCHEDULE CORRECTION PLAN", ""],
        ["Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ["", ""],
        ["TARGET GO-LIVE:", "2026-01-13"],
        ["CURRENT PROJECTION:", "2026-01-30"],
        ["DAYS TO COMPRESS:", "17 days"],
        ["", ""],
        ["CRITICAL PATH:", ""],
        ["  1. IGT Row 24 (SIP Trunks)", "Dec 23 - BLOCKER"],
        ["  2. FPS Row 40 (QA Testing)", "Jan 6"],
        ["  3. Frontier Row 56 (UAT)", "Jan 19"],
        ["  4. Frontier Row 62 (UAT Approval)", "Jan 21"],
        ["  5. Frontier Row 69 (CAB)", "Jan 28"],
        ["  6. FPS Row 72 (Production Deploy)", "Jan 30"],
        ["", ""],
        ["IMMEDIATE ACTIONS:", ""],
        ["  1. Update all baselines", "Shift +6 days (Jan 7 → Jan 13)"],
        ["  2. Escalate IGT dependency", "Can they complete Dec 16?"],
        ["  3. Review Row 24 dependencies", "Can any run in parallel?"],
        ["  4. Compress UAT window", "8 days → 5 days?"],
        ["  5. Compress CAB window", "4 days → 2 days?"],
        ["", ""],
        ["TASKS OVER TARGET:", f"{len([c for c in corrections if c['Schedule Action']])} tasks"],
        ["BASELINE UPDATES NEEDED:", f"{len([c for c in corrections if c['Baseline Action']])} tasks"],
        ["PREDECESSOR REVIEWS:", f"{len([c for c in corrections if c['Predecessor Action']])} tasks"],
    ]

    for row_idx, (label, value) in enumerate(summaries, 1):
        ws2.cell(row=row_idx, column=1, value=label)
        ws2.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=14)

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 35

    # Save
    wb.save('Schedule_Corrections_Plan.xlsx')
    print(f"[OK] Created Schedule_Corrections_Plan.xlsx")

    return corrections


def print_summary(corrections):
    """Print summary to console"""
    print("\n" + "=" * 80)
    print("  SCHEDULE CORRECTIONS SUMMARY")
    print("=" * 80)

    baseline_updates = [c for c in corrections if c['Baseline Action']]
    pred_reviews = [c for c in corrections if c['Predecessor Action']]
    schedule_compress = [c for c in corrections if c['Schedule Action']]

    print(f"\n  Total Tasks: {len(corrections)}")
    print(f"  Baseline Updates Needed: {len(baseline_updates)}")
    print(f"  Predecessor Reviews: {len(pred_reviews)}")
    print(f"  Schedule Compression: {len(schedule_compress)}")

    print(f"\n  TASKS REQUIRING BASELINE UPDATE:")
    for c in baseline_updates:
        print(f"    Row {c['Row']:2}: {c['Task'][:40]} -> {c['New Baseline Finish']}")

    print(f"\n  TASKS BLOCKED BY IGT (Row 24):")
    igt_blocked = [c for c in corrections if 'IGT' in c.get('Predecessor Action', '')]
    for c in igt_blocked:
        print(f"    Row {c['Row']:2}: {c['Assigned To']:10} | {c['Task'][:40]}")

    print(f"\n  TASKS OVER JAN 13 TARGET:")
    for c in schedule_compress[:15]:
        assigned = c['Assigned To'] or 'N/A'
        print(f"    Row {c['Row']:2}: {c['Schedule Action']:12} | {assigned:10} | {c['Task'][:35]}")


def main():
    print("\n" + "=" * 80)
    print("  GENERATING SCHEDULE CORRECTIONS")
    print("=" * 80)

    tasks = load_data()
    print(f"  Loaded {len(tasks)} tasks")

    corrections = calculate_corrections(tasks)
    create_excel_report(corrections)
    print_summary(corrections)

    print("\n" + "=" * 80)
    print("  OUTPUT: Schedule_Corrections_Plan.xlsx")
    print("=" * 80 + "\n")


if __name__ == "__main__":
    main()
