"""
Generate Excel Report with All Corrections for Jan 13 Target
"""

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def load_corrections():
    """Load corrections from JSON"""
    with open('corrections_jan13.json', 'r') as f:
        return json.load(f)


def load_schedule():
    """Load full schedule data"""
    with open('sheet_data_audit.json', 'r') as f:
        return json.load(f)


def create_excel_report():
    """Create comprehensive Excel correction report"""

    corrections = load_corrections()
    schedule = load_schedule()

    # Build lookup
    schedule_by_row = {t['row_number']: t for t in schedule}

    wb = Workbook()

    # ============ Sheet 1: Executive Summary ============
    ws1 = wb.active
    ws1.title = "Executive Summary"

    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    summary_data = [
        ["PHASE 2 AGENTIC VOICE - SCHEDULE CORRECTION PLAN", ""],
        ["Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ["", ""],
        ["TARGET GO-LIVE (APPROVED):", "January 13, 2026"],
        ["ORIGINAL BASELINE:", "January 7, 2026"],
        ["CURRENT PROJECTION:", "January 30, 2026"],
        ["RECALCULATED PROJECTION:", "January 16, 2026"],
        ["", ""],
        ["ROOT CAUSES:", ""],
        ["1. FPS Developer Reassignment", "Nov 25 - Dec 9 (4 working day shift approved)"],
        ["2. IGT SIP Trunks Delay", "Dec 8 -> Dec 23 (2-3 weeks vs 1 week)"],
        ["", ""],
        ["CRITICAL PATH RECALCULATION:", ""],
        ["IGT Staging Complete", "Dec 23 (was Dec 8) +15 days"],
        ["FPS QA Testing Complete", "Jan 5 (was Dec 12) +24 days"],
        ["Frontier UAT Complete", "Jan 12 (was Dec 19) +24 days"],
        ["UAT Approval", "Jan 13 (was Dec 19) +25 days"],
        ["CAB Approval", "Jan 15 (was Jan 6) +9 days"],
        ["Production Go-Live", "Jan 16 (was Jan 7) +9 days"],
        ["", ""],
        ["GAP TO TARGET:", "3 days (Jan 16 vs Jan 13)"],
        ["", ""],
        ["COMPRESSION OPTIONS:", ""],
        ["Option A: Compress UAT 5d -> 3d", "Saves 2 days"],
        ["Option B: Compress CAB 2d -> 1d", "Saves 1 day"],
        ["Option C: Parallel UAT/QA overlap", "Saves 1-2 days"],
        ["", ""],
        ["BASELINE UPDATES REQUIRED:", f"{len([c for c in corrections if c.get('new_baseline_finish')])} tasks"],
        ["TASKS BLOCKED BY IGT:", f"{len([c for c in corrections if 'IGT' in (c.get('notes') or '')])} tasks"],
        ["COMPLETED TASKS (no change):", f"{len([c for c in corrections if c['status'] == 'Complete'])} tasks"],
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws1.cell(row=row_idx, column=1, value=label)
        ws1.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            ws1.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
        if "ROOT CAUSES" in label or "CRITICAL PATH" in label or "COMPRESSION" in label:
            ws1.cell(row=row_idx, column=1).font = Font(bold=True)
            ws1.cell(row=row_idx, column=1).fill = section_fill
            ws1.cell(row=row_idx, column=2).fill = section_fill

    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['B'].width = 45

    # ============ Sheet 2: Baseline Updates ============
    ws2 = wb.create_sheet("Baseline Updates")

    headers = ['Row', 'Task Name', 'Assigned To', 'Status', 'Current Baseline', 'New Baseline', 'Action', 'Notes']

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row_idx = 2
    for c in corrections:
        if c['status'] == 'Complete':
            continue

        ws2.cell(row=row_idx, column=1, value=c['row_number'])
        ws2.cell(row=row_idx, column=2, value=c['task_name'][:50])
        ws2.cell(row=row_idx, column=3, value=c['assigned_to'] or '')
        ws2.cell(row=row_idx, column=4, value=c['status'])
        ws2.cell(row=row_idx, column=5, value=c.get('current_baseline_finish', ''))
        ws2.cell(row=row_idx, column=6, value=c.get('new_baseline_finish', ''))
        ws2.cell(row=row_idx, column=7, value=c.get('baseline_action', ''))
        ws2.cell(row=row_idx, column=8, value=c.get('notes', ''))

        # Color coding
        if 'Jan 7 -> Jan 13' in (c.get('baseline_action') or ''):
            ws2.cell(row=row_idx, column=6).fill = green_fill
        elif 'PROPORTIONAL' in (c.get('baseline_action') or ''):
            ws2.cell(row=row_idx, column=6).fill = yellow_fill

        row_idx += 1

    # Column widths
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 25
    ws2.column_dimensions['H'].width = 35

    ws2.freeze_panes = 'B2'

    # ============ Sheet 3: IGT Blocked Tasks ============
    ws3 = wb.create_sheet("IGT Blocked Tasks")

    headers3 = ['Row', 'Task Name', 'Assigned To', 'Predecessor', 'Current End', 'Recalc End', 'Slip Days']

    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row_idx = 2
    for c in corrections:
        if 'IGT' not in (c.get('notes') or ''):
            continue

        task_data = schedule_by_row.get(c['row_number'], {})
        current_end = task_data.get('End Date', '')[:10] if task_data.get('End Date') else ''
        pred = task_data.get('Predecessors', '')

        ws3.cell(row=row_idx, column=1, value=c['row_number'])
        ws3.cell(row=row_idx, column=2, value=c['task_name'][:45])
        ws3.cell(row=row_idx, column=3, value=c['assigned_to'] or '')
        ws3.cell(row=row_idx, column=4, value=pred)
        ws3.cell(row=row_idx, column=5, value=current_end)
        ws3.cell(row=row_idx, column=6, value='Recalc from Dec 23')
        ws3.cell(row=row_idx, column=7).fill = red_fill

        row_idx += 1

    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 45
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 15
    ws3.column_dimensions['F'].width = 18
    ws3.column_dimensions['G'].width = 10

    # ============ Sheet 4: Full Schedule with Updates ============
    ws4 = wb.create_sheet("Full Schedule")

    headers4 = ['Row', 'Task', 'Assigned', 'Status', 'Health', 'Duration',
                'Current Start', 'Current End', 'Current Baseline', 'New Baseline',
                'Variance', 'Predecessor', 'Notes']

    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    corrections_by_row = {c['row_number']: c for c in corrections}

    row_idx = 2
    for task in schedule:
        corr = corrections_by_row.get(task['row_number'], {})

        start = task.get('Start Date', '')[:10] if task.get('Start Date') else ''
        end = task.get('End Date', '')[:10] if task.get('End Date') else ''
        baseline = task.get('Baseline Finish', '')[:10] if task.get('Baseline Finish') else ''
        new_baseline = corr.get('new_baseline_finish', '')

        ws4.cell(row=row_idx, column=1, value=task['row_number'])
        ws4.cell(row=row_idx, column=2, value=task['Tasks'][:50])
        ws4.cell(row=row_idx, column=3, value=task.get('Assigned To', ''))
        ws4.cell(row=row_idx, column=4, value=task.get('Status', ''))
        ws4.cell(row=row_idx, column=5, value=task.get('Health', ''))
        ws4.cell(row=row_idx, column=6, value=task.get('Duration', ''))
        ws4.cell(row=row_idx, column=7, value=start)
        ws4.cell(row=row_idx, column=8, value=end)
        ws4.cell(row=row_idx, column=9, value=baseline)
        ws4.cell(row=row_idx, column=10, value=new_baseline)
        ws4.cell(row=row_idx, column=11, value=task.get('Variance', ''))
        ws4.cell(row=row_idx, column=12, value=task.get('Predecessors', ''))
        ws4.cell(row=row_idx, column=13, value=task.get('Notes', ''))

        # Health coloring
        health = task.get('Health', '')
        if health == 'Red':
            ws4.cell(row=row_idx, column=5).fill = red_fill
        elif health == 'Yellow':
            ws4.cell(row=row_idx, column=5).fill = yellow_fill
        elif health == 'Green':
            ws4.cell(row=row_idx, column=5).fill = green_fill

        # Highlight new baseline updates
        if new_baseline:
            ws4.cell(row=row_idx, column=10).fill = yellow_fill

        row_idx += 1

    # Column widths
    for col, width in enumerate([5, 45, 10, 12, 8, 8, 12, 12, 12, 12, 10, 15, 35], 1):
        ws4.column_dimensions[chr(64 + col)].width = width

    ws4.freeze_panes = 'C2'

    # ============ Sheet 5: Update Instructions ============
    ws5 = wb.create_sheet("Update Instructions")

    instructions = [
        ["SMARTSHEET UPDATE INSTRUCTIONS", ""],
        ["", ""],
        ["STEP 1: BACKUP THE SHEET", ""],
        ["- Export current sheet to Excel as backup", ""],
        ["- Save with date stamp (e.g., Phase2_Backup_20251209.xlsx)", ""],
        ["", ""],
        ["STEP 2: UPDATE BASELINES", ""],
        ["Option A: Manual Update", ""],
        ["- Open 'Baseline Updates' tab in this file", ""],
        ["- For each row, update Baseline Finish to the New Baseline value", ""],
        ["- Update Baseline Start if shown", ""],
        ["", ""],
        ["Option B: API Update", ""],
        ["- Run: python smartsheet_update_jan13.py", ""],
        ["- Set SHEET_ID in the script first", ""],
        ["- Set SMARTSHEET_ACCESS_TOKEN environment variable", ""],
        ["", ""],
        ["STEP 3: VERIFY CRITICAL PATH", ""],
        ["- Check Row 24 (IGT SIP Trunks) shows Dec 23 end date", ""],
        ["- Verify dependent tasks cascade correctly from Dec 23", ""],
        ["- Confirm FPS QA -> Frontier UAT -> CAB chain is correct", ""],
        ["", ""],
        ["STEP 4: ADD NOTES", ""],
        ["Row 25 (IGT Signal API):", "Already has 12/05 note about delay"],
        ["Row 28 (FPS Knowledgebase):", "Add: Developer reassigned Nov 25 - Dec 9"],
        ["Row 64 (Production Deployment):", "Add: Baseline updated to Jan 13 per approval"],
        ["", ""],
        ["STEP 5: RECALCULATE", ""],
        ["- Smartsheet should auto-recalculate variance", ""],
        ["- Health indicators will update", ""],
        ["- Verify final Production Deployment date", ""],
        ["", ""],
        ["EXPECTED RESULTS:", ""],
        ["- Variance will show against Jan 13 baseline", ""],
        ["- Production Deployment projection: ~Jan 16", ""],
        ["- Gap to target: 3 days", ""],
    ]

    for row_idx, (label, value) in enumerate(instructions, 1):
        ws5.cell(row=row_idx, column=1, value=label)
        ws5.cell(row=row_idx, column=2, value=value)
        if "STEP" in label or label.startswith("SMARTSHEET") or "EXPECTED" in label:
            ws5.cell(row=row_idx, column=1).font = Font(bold=True)
            ws5.cell(row=row_idx, column=1).fill = section_fill

    ws5.column_dimensions['A'].width = 45
    ws5.column_dimensions['B'].width = 50

    # Save
    filename = 'Schedule_Corrections_Jan13.xlsx'
    wb.save(filename)
    print(f"[OK] Created {filename}")

    return filename


if __name__ == "__main__":
    filename = create_excel_report()
    print(f"\nExcel report saved: {filename}")
